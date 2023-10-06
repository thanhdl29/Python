# import openpyxl and tkinter modules
from openpyxl import *
from tkinter import *

# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook('D:\\Workspace\\LapTrinh_PyThon\\ThongTinDangKyHocPhan.xlsx')

# create the sheet object
sheet = wb.active


def excel():
	# resize the width of columns in
	# excel spreadsheet
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50
	sheet.column_dimensions['H'].width = 50
    
	# write given data to an excel spreadsheet
	# at particular location
	sheet.cell(row=1, column=1).value = "Mã số sinh viên"
	sheet.cell(row=1, column=2).value = "Họ tên"
	sheet.cell(row=1, column=3).value = "ngày sinh"
	sheet.cell(row=1, column=4).value = "Email"
	sheet.cell(row=1, column=5).value = "Số điện thoại"
	sheet.cell(row=1, column=6).value = "Học kỳ"
	sheet.cell(row=1, column=7).value = "Năm học"
	sheet.cell(row=1, column=8).value = "Chọn môn học"
 


# Function to set focus (cursor)
def focus1(event):
	# set focus on the course_field box
	Mssv_field.focus_set()


# Function to set focus
def focus2(event):
	# set focus on the sem_field box
	HoTen_field.focus_set()


# Function to set focus
def focus3(event):
	# set focus on the form_no_field box
	NgaySinh_field.focus_set()


# Function to set focus
def focus4(event):
	# set focus on the contact_no_field box
	Email_field.focus_set()


# Function to set focus
def focus5(event):
	# set focus on the email_id_field box
	SoDienThoai_field.focus_set()


# Function to set focus
def focus6(event):
	# set focus on the address_field box
	HocKy_field.focus_set()
 
def focus7(event):
	# set focus on the address_field box
	NamHoc_field.focus_set()
def focus8(event):
	# set focus on the address_field box
	ChonMonHoc_field.focus_set()

# Function for clearing the
# contents of text entry boxes
def clear():
	
	# clear the content of text entry box
	Mssv_field.delete(0, END)
	HoTen_field.delete(0, END)
	NgaySinh_field.delete(0, END)
	Email_field.delete(0, END)
	SoDienThoai_field.delete(0, END)
	HocKy_field.delete(0, END)
	NamHoc_field.delete(0, END)
	ChonMonHoc_field.delete(0, END)


# Function to take data from GUI
# window and write to an excel file
def insert():
	
	# if user not fill any entry
	# then print "empty input"
	if (Mssv_field.get() == "" and
		HoTen_field.get() == "" and
		NgaySinh_field.get() == "" and
		Email_field.get() == "" and
		SoDienThoai_field.get() == "" and
		HocKy_field.get() == "" and
		NamHoc_field.get() == ""and
        ChonMonHoc_field.get() == ""):
			
		print("empty input")

	else:

		# assigning the max row and max column
		# value upto which data is written
		# in an excel sheet to the variable
		current_row = sheet.max_row
		current_column = sheet.max_column

		# get method returns current text
		# as string which we write into
		# excel spreadsheet at particular location
		sheet.cell(row=current_row + 1, column=1).value = Mssv_field.get()
		sheet.cell(row=current_row + 1, column=2).value = HoTen_field.get()
		sheet.cell(row=current_row + 1, column=3).value = NgaySinh_field.get()
		sheet.cell(row=current_row + 1, column=4).value = Email_field.get()
		sheet.cell(row=current_row + 1, column=5).value = SoDienThoai_field.get()
		sheet.cell(row=current_row + 1, column=6).value = HocKy_field.get()
		sheet.cell(row=current_row + 1, column=7).value = NamHoc_field.get()
		sheet.cell(row=current_row + 1, column=8).value = ChonMonHoc_field.get()
        
		# save the file
		wb.save('D:\\Workspace\\LapTrinh_PyThon\\ThongTinDangKyHocPhan.xlsx')

		# set focus on the name_field box
		Mssv_field.focus_set()

		# call the clear() function
		clear()


# Driver code
if __name__ == "__main__":
	
	# create a GUI window
	root = Tk()

	# set the background colour of GUI window
	root.configure(background='light green')

	# set the title of GUI window
	root.title("registration form")

	# set the configuration of GUI window
	root.geometry("500x300")

	excel()

	# create a Form label
	heading = Label(root, text="THÔNG TIN ĐĂNG KÝ HỌC PHẦN", bg="light green")

	# create a Name label
	Mssv = Label(root, text="Mã số sinh viên", bg="light green")

	# create a Course label
	HoTen = Label(root, text="Họ Tên", bg="light green")

	# create a Semester label
	NgaySinh = Label(root, text="Ngay sinh", bg="light green")

	# create a Form No. label
	Email = Label(root, text="Email", bg="light green")

	# create a Contact No. label
	SoDienThoai = Label(root, text="Số điện thoại", bg="light green")

	# create a Email id label
	HocKy = Label(root, text="Học kỳ", bg="light green")

	# create a address label
	NamHoc = Label(root, text="Năm học", bg="light green")
 
	ChonMonHoc = Checkbutton(root, text="Chọn môn học", bg="light green")
	    

	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
	heading.grid(row=0, column=1)
	Mssv.grid(row=1, column=0)
	HoTen.grid(row=2, column=0)
	NgaySinh.grid(row=3, column=0)
	Email.grid(row=4, column=0)
	SoDienThoai.grid(row=5, column=0)
	HocKy.grid(row=6, column=0)
	NamHoc.grid(row=7, column=0)
	ChonMonHoc.grid(row=8, column=0)

	# create a text entry box
	# for typing the information
	Mssv_field = Entry(root)
	HoTen_field = Entry(root)
	NgaySinh_field = Entry(root)
	Email_field = Entry(root)
	SoDienThoai_field = Entry(root)
	HocKy_field = Entry(root)
	NamHoc_field = Entry(root)
	ChonMonHoc_field = Entry(root)

	# bind method of widget is used for
	# the binding the function with the events

	# whenever the enter key is pressed
	# then call the focus1 function
	Mssv_field.bind("<Return>", focus1)

	# whenever the enter key is pressed
	# then call the focus2 function
	HoTen_field.bind("<Return>", focus2)

	# whenever the enter key is pressed
	# then call the focus3 function
	NgaySinh.bind("<Return>", focus3)

	# whenever the enter key is pressed
	# then call the focus4 function
	Email.bind("<Return>", focus4)

	# whenever the enter key is pressed
	# then call the focus5 function
	SoDienThoai.bind("<Return>", focus5)

	# whenever the enter key is pressed
	# then call the focus6 function
	HocKy.bind("<Return>", focus6)
	NamHoc.bind("<Return>", focus7)
	ChonMonHoc.bind("<Return>", focus8)

	# grid method is used for placing
	# the widgets at respective positions
	# in table like structure .
	Mssv_field.grid(row=1, column=1, ipadx="100")
	HoTen_field.grid(row=2, column=1, ipadx="100")
	NgaySinh.grid(row=3, column=1, ipadx="100")
	Email_field.grid(row=4, column=1, ipadx="100")
	SoDienThoai_field.grid(row=5, column=1, ipadx="100")
	HocKy_field.grid(row=6, column=1, ipadx="100")
	NamHoc_field.grid(row=7, column=1, ipadx="100")
	ChonMonHoc_field.grid(row=8, column=1, ipadx="100")

	# call excel function
	excel()

	# create a Submit Button and place into the root window
	submit = Button(root, text="Submit", fg="Black",
							bg="Red", command=insert)
	submit.grid(row=9, column=1)

	# start the GUI
	root.mainloop()
